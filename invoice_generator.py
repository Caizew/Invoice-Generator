from abc import ABC, abstractmethod
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from datetime import datetime


class InvoiceGenerator(ABC):

    def __init__(self, client_name, items):
        self.client_name = client_name
        self.items = items

    def calculate_total(self):
        total = 0
        for item in self.items:
            if "price" in item and type(item["price"]) in [int, float]:
                total = total + item["price"]
        return total

    @abstractmethod
    def generate_invoice(self):
        pass


class PDF_invoiceGenerator(InvoiceGenerator):
    def __init__(self, client_name, items):
        super().__init__(client_name, items)

    def generate_invoice(self):
        pdf = canvas.Canvas(f"invoice_{self.client_name}.pdf")
        pdf.drawString(100, 800, "Invoice")
        pdf.drawString(100, 780, f"Client: {self.client_name}")
        starty = 760
        for item in self.items:
            pdf.drawString(100, starty, f"{item['name']} - ${item.get('price', 0)}")
            starty = starty - 20
        pdf.drawString(100, starty - 20, f"Generated on: {datetime.now()}")
        pdf.save()


class Excell_Invoice_Generator(InvoiceGenerator):
    def __init__(self, client_name, items):
        super().__init__(client_name, items)

    def generate_invoice(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Invoice"
        ws["A2"] = f"Client: {self.client_name}"
        ws["A4"] = "item name"
        ws["B4"] = "Price"
        row = 5
        for item in self.items:
            ws[f"A{row}"] = item["name"]
            ws[f"B{row}"] = item.get("price", 0)
            row = row + 1
        total = self.calculate_total()
        ws[f"A{row+1}"] = "Total"
        ws[f"B{row+1}"] = total
        ws[f"A{row+2}"] = f"Generated on: {datetime.now()}"

        wb.save(f"invoice_{self.client_name}.xlsx")


class HTMLInvoiceGenerator(InvoiceGenerator):
    def __init__(self, client_name, items):
        super().__init__(client_name, items)

    def generate_invoice(self):
        # Create the HTML file
        with open(f"invoice_{self.client_name}.html", "w") as file:
            file.write(
                """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', sans-serif; background: #f5f7fa; padding: 40px 20px; }
        .container { max-width: 900px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 10px 40px rgba(0,0,0,0.1); overflow: hidden; }
        .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 40px; }
        .header h1 { font-size: 32px; margin-bottom: 4px; font-weight: 700; }
        .content { padding: 40px; }
        .client { margin-bottom: 40px; padding-bottom: 40px; border-bottom: 1px solid #e9ecef; }
        .client p { font-size: 16px; color: #2d3748; margin: 4px 0; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 30px; }
        thead { background: #f8f9fa; border-bottom: 2px solid #667eea; }
        th { padding: 14px; text-align: left; font-size: 12px; font-weight: 600; color: #667eea; text-transform: uppercase; letter-spacing: 0.5px; }
        td { padding: 14px; border-bottom: 1px solid #e9ecef; font-size: 15px; color: #2d3748; }
        tbody tr:hover { background: #f8f9fa; }
        tbody tr:last-child td { border-bottom: none; }
        .total-section { display: flex; justify-content: flex-end; margin-top: 20px; }
        .total-box { text-align: right; padding: 20px; background: #f8f9fa; border-radius: 8px; border-left: 4px solid #667eea; }
        .total-label { font-size: 12px; color: #6c757d; text-transform: uppercase; letter-spacing: 1px; font-weight: 600; margin-bottom: 8px; display: block; }
        .total-value { font-size: 28px; font-weight: 700; color: #667eea; }
        .footer { background: #f8f9fa; padding: 20px 40px; border-top: 1px solid #e9ecef; font-size: 13px; color: #6c757d; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Invoice</h1>
        </div>
        <div class="content">
            <div class="client">"""
            )
            file.write(f"<p><strong>Client: {self.client_name}</strong></p>")
            file.write(
                """</div>
            <table border='0'>
                <thead>
                    <tr><th>Item Name</th><th style="text-align: right;">Price ($)</th></tr>
                </thead>
                <tbody>"""
            )
            # Add items
            for item in self.items:
                file.write(
                    f"<tr><td>{item['name']}</td><td style='text-align: right;'>${item.get('price', 0):.2f}</td></tr>"
                )
            # Add total and date
            total = self.calculate_total()
            file.write(
                f"""</tbody>
            </table>
            <div class="total-section">
                <div class="total-box">
                    <span class="total-label">Total</span>
                    <span class="total-value">${total:.2f}</span>
                </div>
            </div>
        </div>
        <div class="footer">Generated on: {datetime.now()}</div>
    </div>
</body>
</html>"""
            )


items = [
    {"name": "Spoon", "price": 3.59},
    {"name": "Car", "price": 9000},
    {"name": "Computer", "price": 1000},
]

pdf_gen = PDF_invoiceGenerator("Palonchi", items)
pdf_gen.generate_invoice()

excel_gen = Excell_Invoice_Generator("Palonchi", items)
excel_gen.generate_invoice()

html_gen = HTMLInvoiceGenerator("PAlonchi", items)
html_gen.generate_invoice()
